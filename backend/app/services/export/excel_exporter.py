from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill

SEVERITY_FILLS = {
    "CRITICAL": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "HIGH": PatternFill(fill_type="solid", fgColor="F4B183"),
    "MEDIUM": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "LOW": PatternFill(fill_type="solid", fgColor="E2F0D9"),
}


class ExcelExporter:
    def build_validation_report(self, *, batch, invoices, errors, reconciliation_results) -> bytes:
        workbook = Workbook()
        ws_invoices = workbook.active
        ws_invoices.title = "All Invoices"
        ws_invoices.append(["Invoice Number", "State", "Supplier GSTIN", "Total Value"])
        for invoice in invoices:
            ws_invoices.append([invoice.invoice_number, invoice.state.value, invoice.supplier_gstin, str(invoice.total_invoice_value)])
        ws_errors = workbook.create_sheet("Errors Only")
        ws_errors.append(["Invoice ID", "Rule", "Severity", "Message"])
        for error in errors:
            ws_errors.append([str(error.invoice_id), error.rule_id, error.severity.value, error.message])
            for cell in ws_errors[ws_errors.max_row]:
                cell.fill = SEVERITY_FILLS[error.severity.value]
        ws_recon = workbook.create_sheet("GSTR-2B Reconciliation")
        ws_recon.append(["Invoice ID", "Match Type", "Confidence"])
        for result in reconciliation_results:
            ws_recon.append([str(result.invoice_id), result.match_type.value, result.confidence_score])
        ws_summary = workbook.create_sheet("Summary")
        ws_summary.append(["Batch ID", str(batch.id)])
        ws_summary.append(["Invoice Count", len(invoices)])
        ws_summary.append(["Error Count", len(errors)])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
